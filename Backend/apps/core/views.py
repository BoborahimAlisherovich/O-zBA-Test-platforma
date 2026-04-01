from io import BytesIO

from django.db import connection
from django.utils import timezone
from datetime import timedelta
from rest_framework import parsers, status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from openpyxl import load_workbook

from apps.accounts.models import User
from .models import Group, Module, Question, ResultArchiveFolder, SiteSetting, Subject, TestAttempt, TestResult
from .permissions import IsAdminOnly, IsParticipantOnly
from .serializers import (
    GroupSerializer,
    ModuleSerializer,
    ParticipantSnapshotSerializer,
    QuestionSerializer,
    ResultArchiveFolderSerializer,
    SnapshotSerializer,
    SiteSettingSerializer,
    TestAttemptSerializer,
    SubjectSerializer,
    TestResultSerializer,
)
from .services import pick_questions_for_module


def _to_bool(v, default=False):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in {"1", "true", "yes", "on"}
    return bool(v)


def _to_int(v, default=None):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _normalize_answers(raw_answers, allowed_question_ids):
    normalized = {}
    allowed = {str(qid) for qid in allowed_question_ids}
    for key, value in (raw_answers or {}).items():
        normalized_key = str(key)
        if normalized_key not in allowed:
            continue
        parsed_value = _to_int(value, None)
        if parsed_value is None:
            continue
        normalized[normalized_key] = parsed_value
    return normalized


def _normalize_remaining_seconds(value, fallback):
    parsed = _to_int(value, fallback)
    if parsed is None:
        return max(0, int(fallback or 0))
    return max(0, int(parsed))


def _get_site_settings():
    site_settings, _ = SiteSetting.objects.get_or_create(pk=1)
    return site_settings


def _get_demo_max_attempts():
    return max(1, int(_get_site_settings().demo_max_attempts or 1))


def _validate_module_access(user, module):
    if user.group_id is None or not module.groups.filter(id=user.group_id).exists():
        return Response({"detail": "Siz bu testga biriktirilmagansiz"}, status=status.HTTP_403_FORBIDDEN)
    if module.is_demo and TestResult.objects.filter(participant=user, module=module).count() >= _get_demo_max_attempts():
        return Response({"detail": "Sizda limit tugadi"}, status=status.HTTP_400_BAD_REQUEST)
    if not module.is_demo and TestResult.objects.filter(participant=user, module=module).exists():
        return Response({"detail": "Bu test allaqachon topshirilgan"}, status=status.HTTP_400_BAD_REQUEST)
    return None


def _finalize_attempt(attempt, submitted_answers=None, submitted_time_taken=None):
    if hasattr(attempt, "result"):
        return attempt.result

    question_payload = attempt.question_payload or []
    question_ids = [row.get("id") for row in question_payload if row.get("id") is not None]
    questions = Question.objects.filter(id__in=question_ids)
    question_map = {str(q.id): q for q in questions}
    answers = _normalize_answers(submitted_answers or attempt.answers or {}, question_map.keys())

    correct = 0
    for qid, question in question_map.items():
        chosen = answers.get(str(qid))
        if chosen is not None and int(chosen) == question.correct_index:
            correct += 1

    total = len(question_payload)
    score = correct * attempt.module.points_per_answer
    is_passed = score >= attempt.module.passing_score
    if submitted_time_taken is not None:
        time_taken = _to_int(submitted_time_taken, attempt.module.duration_minutes * 60)
    else:
        total_allocated = attempt.module.duration_minutes * 60
        time_taken = min(total_allocated, max(0, total_allocated - int(attempt.remaining_seconds or 0)))

    result = TestResult.objects.create(
        participant=attempt.participant,
        module=attempt.module,
        group=attempt.group,
        attempt=attempt,
        correct_answers=correct,
        total_questions=total,
        score=score,
        is_passed=is_passed,
        time_taken=time_taken,
    )
    attempt.answers = answers
    attempt.remaining_seconds = 0
    attempt.completed_at = timezone.now()
    attempt.save(update_fields=["answers", "remaining_seconds", "completed_at", "updated_at"])
    return result


def _expire_stale_attempts(user):
    expired_attempts = TestAttempt.objects.select_related("module", "participant", "group").filter(
        participant=user,
        completed_at__isnull=True,
        remaining_seconds__lte=0,
    )
    for attempt in expired_attempts:
        _finalize_attempt(attempt)


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.prefetch_related("modules").all().order_by("-id")
    serializer_class = GroupSerializer
    permission_classes = [IsAdminOnly]


class SubjectViewSet(viewsets.ModelViewSet):
    serializer_class = SubjectSerializer
    permission_classes = [IsAdminOnly]

    def get_queryset(self):
        is_demo = self.request.query_params.get("is_demo")
        qs = Subject.objects.all().order_by("-id")
        if is_demo is not None:
            qs = qs.filter(is_demo=str(is_demo).lower() in {"1", "true", "yes"})
        return qs

    def perform_create(self, serializer):
        is_demo = self.request.query_params.get("is_demo")
        if is_demo is None:
            serializer.save()
            return
        serializer.save(is_demo=str(is_demo).lower() in {"1", "true", "yes"})


class ModuleViewSet(viewsets.ModelViewSet):
    serializer_class = ModuleSerializer
    permission_classes = [IsAdminOnly]

    def get_queryset(self):
        is_demo = self.request.query_params.get("is_demo")
        qs = Module.objects.prefetch_related("groups", "subject_configs").all().order_by("-id")
        if is_demo is not None:
            qs = qs.filter(is_demo=str(is_demo).lower() in {"1", "true", "yes"})
        return qs

    def perform_create(self, serializer):
        is_demo = self.request.query_params.get("is_demo")
        if is_demo is None:
            serializer.save()
            return
        serializer.save(is_demo=str(is_demo).lower() in {"1", "true", "yes"})


class QuestionViewSet(viewsets.ModelViewSet):
    serializer_class = QuestionSerializer
    permission_classes = [IsAdminOnly]

    def get_queryset(self):
        qs = Question.objects.select_related("subject").all().order_by("-id")
        is_demo = self.request.query_params.get("is_demo")
        subject_id = self.request.query_params.get("subject_id")
        if is_demo is not None:
            qs = qs.filter(subject__is_demo=str(is_demo).lower() in {"1", "true", "yes"})
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        return qs


class TestResultViewSet(viewsets.ModelViewSet):
    serializer_class = TestResultSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "delete", "patch", "head", "options"]

    def get_queryset(self):
        qs = TestResult.objects.select_related("participant", "module", "group").all()
        user = self.request.user
        if user.role in {"ADMIN", "MANAGER"}:
            return qs
        return qs.filter(participant=user)

    def get_permissions(self):
        if self.action in {"destroy", "partial_update"}:
            return [IsAdminOnly()]
        return [permission() for permission in self.permission_classes]


class ResultArchiveFolderViewSet(viewsets.ModelViewSet):
    queryset = ResultArchiveFolder.objects.prefetch_related("results").all().order_by("-created_at", "-id")
    serializer_class = ResultArchiveFolderSerializer
    permission_classes = [IsAdminOnly]
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]


class SiteSettingViewSet(viewsets.ModelViewSet):
    queryset = SiteSetting.objects.all().order_by("id")
    serializer_class = SiteSettingSerializer
    permission_classes = [IsAdminOnly]
    parser_classes = [parsers.MultiPartParser, parsers.FormParser, parsers.JSONParser]

    def list(self, request, *args, **kwargs):
        site_settings = _get_site_settings()
        serializer = self.get_serializer(site_settings)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        site_settings = _get_site_settings()
        serializer = self.get_serializer(site_settings, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        site_settings = _get_site_settings()
        serializer = self.get_serializer(site_settings)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        site_settings = _get_site_settings()
        serializer = self.get_serializer(site_settings, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


def _build_snapshot_payload(user):
    if user.role in {"ADMIN", "MANAGER"}:
        users = User.objects.select_related("group").filter(is_archived=False).order_by("-id")
        archived_users = User.objects.select_related("group").filter(is_archived=True).order_by("-id")
        groups = Group.objects.prefetch_related("modules").all().order_by("-id")
        subjects = Subject.objects.filter(is_demo=False).order_by("-id")
        demo_subjects = Subject.objects.filter(is_demo=True).order_by("-id")
        modules = Module.objects.filter(is_demo=False).prefetch_related("groups", "subject_configs").order_by("-id")
        demo_modules = Module.objects.filter(is_demo=True).prefetch_related("groups", "subject_configs").order_by("-id")
        questions = Question.objects.filter(subject__is_demo=False).order_by("-id")
        demo_questions = Question.objects.filter(subject__is_demo=True).order_by("-id")
        results = TestResult.objects.filter(module__is_demo=False, archive_folder__isnull=True).order_by("-date")
        archived_results = TestResult.objects.filter(module__is_demo=False, archive_folder__isnull=False).order_by("-date")
        demo_results = TestResult.objects.filter(module__is_demo=True).order_by("-date")
        result_archive_folders = ResultArchiveFolder.objects.prefetch_related("results").all().order_by("-created_at", "-id")
    else:
        group_id = user.group_id
        users = User.objects.filter(id=user.id)
        archived_users = User.objects.none()
        groups = Group.objects.prefetch_related("modules").filter(id=group_id) if group_id else Group.objects.none()
        modules = Module.objects.filter(is_demo=False, is_active=True, groups__id=group_id).prefetch_related("groups", "subject_configs").distinct()
        demo_modules = Module.objects.filter(is_demo=True, is_active=True, groups__id=group_id).prefetch_related("groups", "subject_configs").distinct()
        subject_ids = set()
        for m in list(modules) + list(demo_modules):
            subject_ids.update(m.subject_configs.values_list("subject_id", flat=True))
        subjects = Subject.objects.filter(id__in=subject_ids, is_demo=False).order_by("-id")
        demo_subjects = Subject.objects.filter(id__in=subject_ids, is_demo=True).order_by("-id")
        questions = Question.objects.filter(subject__in=subjects).order_by("-id")
        demo_questions = Question.objects.filter(subject__in=demo_subjects).order_by("-id")
        results = TestResult.objects.filter(participant=user, module__is_demo=False).order_by("-date")
        archived_results = TestResult.objects.none()
        demo_results = TestResult.objects.filter(participant=user, module__is_demo=True).order_by("-date")
        result_archive_folders = ResultArchiveFolder.objects.none()

    return {
        "users": users,
        "archivedUsers": archived_users,
        "groups": groups,
        "subjects": subjects,
        "modules": modules,
        "questions": questions,
        "results": results,
        "archivedResults": archived_results,
        "resultArchiveFolders": result_archive_folders,
        "demoSubjects": demo_subjects,
        "demoModules": demo_modules,
        "demoQuestions": demo_questions,
        "demoResults": demo_results,
        "siteSettings": _get_site_settings(),
    }


def _normalize_header(value):
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def _sheet_rows(uploaded_file):
    workbook = load_workbook(filename=BytesIO(uploaded_file.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(cell or "").strip() for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        data_rows.append({headers[idx]: row[idx] for idx in range(len(headers))})
    return headers, data_rows


def _get_value_by_aliases(row, aliases):
    normalized_row = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized_row.get(_normalize_header(alias))
        if value not in (None, ""):
            return value
    return ""


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsParticipantOnly])
def available_tests_view(request):
    user = request.user
    _expire_stale_attempts(user)
    group_id = user.group_id
    if not group_id:
        return Response({"main": [], "demo": []})

    main_modules = Module.objects.filter(is_demo=False, is_active=True, groups__id=group_id).distinct()
    demo_modules = Module.objects.filter(is_demo=True, is_active=True, groups__id=group_id).distinct()
    taken_main = set(TestResult.objects.filter(participant=user, module__is_demo=False).values_list("module_id", flat=True))
    unfinished_modules = set(
        TestAttempt.objects.filter(participant=user, completed_at__isnull=True).values_list("module_id", flat=True)
    )

    return Response(
        {
            "main": [
                {
                    "id": m.id,
                    "name": m.name,
                    "alreadyTaken": m.id in taken_main,
                    "hasUnfinishedAttempt": m.id in unfinished_modules,
                    "settings": {
                        "pointsPerAnswer": m.points_per_answer,
                        "durationMinutes": m.duration_minutes,
                        "passingScore": m.passing_score,
                        "randomize": m.randomize,
                        "isActive": m.is_active,
                    },
                }
                for m in main_modules
            ],
            "demo": [
                {
                    "id": m.id,
                    "name": m.name,
                    "hasUnfinishedAttempt": m.id in unfinished_modules,
                    "settings": {
                        "pointsPerAnswer": m.points_per_answer,
                        "durationMinutes": m.duration_minutes,
                        "passingScore": m.passing_score,
                        "randomize": m.randomize,
                        "isActive": m.is_active,
                    },
                }
                for m in demo_modules
            ],
        }
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsParticipantOnly])
def start_test_view(request):
    _expire_stale_attempts(request.user)
    module_id = request.data.get("moduleId")
    if not module_id:
        return Response({"detail": "moduleId kerak"}, status=status.HTTP_400_BAD_REQUEST)

    module = Module.objects.prefetch_related("subject_configs").filter(id=module_id, is_active=True).first()
    if not module:
        return Response({"detail": "Test topilmadi"}, status=status.HTTP_404_NOT_FOUND)

    access_error = _validate_module_access(request.user, module)
    if access_error:
        return access_error

    existing_attempt = TestAttempt.objects.select_related("module", "participant", "group").filter(
        participant=request.user,
        module=module,
        completed_at__isnull=True,
    ).order_by("-updated_at").first()
    if existing_attempt:
        return Response(TestAttemptSerializer(existing_attempt).data)

    attempt = TestAttempt.objects.create(
        participant=request.user,
        module=module,
        group=request.user.group,
        question_payload=pick_questions_for_module(module),
        answers={},
        current_question_index=0,
        started_at=timezone.now(),
        expires_at=timezone.now() + timedelta(minutes=module.duration_minutes),
        remaining_seconds=module.duration_minutes * 60,
    )
    return Response(TestAttemptSerializer(attempt).data, status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsParticipantOnly])
def active_attempts_view(request):
    _expire_stale_attempts(request.user)
    attempts = TestAttempt.objects.select_related("module", "participant", "group").filter(
        participant=request.user,
        completed_at__isnull=True,
    ).order_by("-updated_at")
    return Response({"attempts": TestAttemptSerializer(attempts, many=True).data})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsParticipantOnly])
def save_test_progress_view(request):
    _expire_stale_attempts(request.user)
    attempt_id = request.data.get("attemptId")
    answers = request.data.get("answers", {})
    current_question_index = _to_int(request.data.get("currentQuestionIndex"), 0)
    remaining_seconds = request.data.get("timeRemaining")

    if not attempt_id:
        return Response({"detail": "attemptId kerak"}, status=status.HTTP_400_BAD_REQUEST)

    attempt = TestAttempt.objects.select_related("module", "participant", "group").filter(
        id=attempt_id,
        participant=request.user,
    ).first()
    if not attempt:
        return Response({"detail": "Test urinish topilmadi"}, status=status.HTTP_404_NOT_FOUND)
    if attempt.completed_at:
        return Response({"detail": "Bu urinish allaqachon yakunlangan"}, status=status.HTTP_400_BAD_REQUEST)
    if attempt.remaining_seconds <= 0:
        _finalize_attempt(attempt)
        return Response({"detail": "Test vaqti tugagan"}, status=status.HTTP_409_CONFLICT)

    question_ids = [row.get("id") for row in (attempt.question_payload or []) if row.get("id") is not None]
    normalized_answers = attempt.answers or {}
    normalized_answers.update(_normalize_answers(answers, question_ids))
    max_index = max(0, len(question_ids) - 1)
    attempt.answers = normalized_answers
    attempt.current_question_index = min(max(current_question_index or 0, 0), max_index) if question_ids else 0
    if remaining_seconds is not None:
        attempt.remaining_seconds = min(
            int(attempt.remaining_seconds or 0),
            _normalize_remaining_seconds(remaining_seconds, attempt.remaining_seconds or 0),
        )
    attempt.save(update_fields=["answers", "current_question_index", "remaining_seconds", "updated_at"])
    return Response(TestAttemptSerializer(attempt).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsParticipantOnly])
def submit_test_view(request):
    _expire_stale_attempts(request.user)
    attempt_id = request.data.get("attemptId")
    module_id = request.data.get("moduleId")
    answers = request.data.get("answers", {})
    time_taken = request.data.get("timeTaken")
    remaining_seconds = request.data.get("timeRemaining")

    if attempt_id:
        attempt = TestAttempt.objects.select_related("module", "participant", "group").filter(
            id=attempt_id,
            participant=request.user,
        ).first()
        if not attempt:
            return Response({"detail": "Test urinish topilmadi"}, status=status.HTTP_404_NOT_FOUND)
        if attempt.completed_at and hasattr(attempt, "result"):
            return Response(TestResultSerializer(attempt.result).data)

        question_ids = [row.get("id") for row in (attempt.question_payload or []) if row.get("id") is not None]
        merged_answers = attempt.answers or {}
        merged_answers.update(_normalize_answers(answers if isinstance(answers, dict) else {}, question_ids))
        if not merged_answers:
            return Response({"detail": "Javoblar topilmadi"}, status=status.HTTP_400_BAD_REQUEST)

        attempt.answers = merged_answers
        if remaining_seconds is not None:
            attempt.remaining_seconds = min(
                int(attempt.remaining_seconds or 0),
                _normalize_remaining_seconds(remaining_seconds, attempt.remaining_seconds or 0),
            )
        attempt.current_question_index = min(
            max(_to_int(request.data.get("currentQuestionIndex"), attempt.current_question_index), 0),
            max(0, len(question_ids) - 1),
        ) if question_ids else 0
        attempt.save(update_fields=["answers", "current_question_index", "remaining_seconds", "updated_at"])
        result = _finalize_attempt(attempt, submitted_answers=merged_answers, submitted_time_taken=time_taken)
        return Response(TestResultSerializer(result).data, status=status.HTTP_201_CREATED)

    if not module_id or not isinstance(answers, dict) or not answers:
        return Response({"detail": "moduleId va answers kerak"}, status=status.HTTP_400_BAD_REQUEST)

    module = Module.objects.filter(id=module_id, is_active=True).first()
    if not module:
        return Response({"detail": "Test topilmadi"}, status=status.HTTP_404_NOT_FOUND)

    access_error = _validate_module_access(request.user, module)
    if access_error:
        return access_error

    try:
        question_ids = [int(k) for k in answers.keys()]
    except ValueError:
        return Response({"detail": "answers keylari savol ID bo'lishi kerak"}, status=status.HTTP_400_BAD_REQUEST)

    allowed_subject_ids = module.subject_configs.values_list("subject_id", flat=True)
    questions = Question.objects.filter(id__in=question_ids, subject_id__in=allowed_subject_ids)
    if questions.count() != len(question_ids):
        return Response({"detail": "Ba'zi savollar topilmadi"}, status=status.HTTP_400_BAD_REQUEST)

    correct = 0
    for q in questions:
        chosen = answers.get(str(q.id), answers.get(q.id))
        if chosen is not None and int(chosen) == q.correct_index:
            correct += 1

    total = len(question_ids)
    score = correct * module.points_per_answer
    is_passed = score >= module.passing_score

    result = TestResult.objects.create(
        participant=request.user,
        module=module,
        group=request.user.group,
        correct_answers=correct,
        total_questions=total,
        score=score,
        is_passed=is_passed,
        time_taken=int(time_taken) if time_taken is not None else None,
    )

    return Response(TestResultSerializer(result).data, status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([AllowAny])
def health_view(request):
    db_ok = True
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
    except Exception:
        db_ok = False

    payload = {
        "status": "ok" if db_ok else "degraded",
        "database": "ok" if db_ok else "error",
        "time": timezone.now(),
    }
    http_status = status.HTTP_200_OK if db_ok else status.HTTP_503_SERVICE_UNAVAILABLE
    return Response(payload, status=http_status)


@api_view(["GET"])
@permission_classes([AllowAny])
def site_settings_view(request):
    site_settings = _get_site_settings()
    return Response(SiteSettingSerializer(site_settings, context={"request": request}).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminOnly])
def import_users_view(request):
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"detail": "Excel fayl yuborilmadi"}, status=status.HTTP_400_BAD_REQUEST)

    _, rows = _sheet_rows(uploaded_file)
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        full_name = str(_get_value_by_aliases(row, ["F.I.SH", "FISH", "fullname", "full_name"])).strip()
        username = str(_get_value_by_aliases(row, ["Login", "username", "user"])).strip()
        password = str(_get_value_by_aliases(row, ["Parol", "password", "paroli"])).strip()
        workplace = str(_get_value_by_aliases(row, ["Asosiy ish joyi", "workplace", "ishjoyi"])).strip()
        role = str(_get_value_by_aliases(row, ["Rol", "role"])).strip() or "TINGLOVCHI"
        group_value = str(_get_value_by_aliases(row, ["Guruh", "group", "groupid"])).strip()

        if not full_name or not username:
            skipped += 1
            errors.append(f"{index}-qator: F.I.SH yoki login bo'sh.")
            continue

        normalized_role = role.upper()
        if normalized_role not in {"ADMIN", "MANAGER", "TINGLOVCHI"}:
            normalized_role = "TINGLOVCHI"

        group = None
        if group_value:
            group = Group.objects.filter(name__iexact=group_value).first()
            if not group and group_value.isdigit():
                group = Group.objects.filter(id=int(group_value)).first()

        user = User.objects.filter(username=username).first()
        if user:
            user.full_name = full_name
            user.workplace = workplace
            user.role = normalized_role
            user.group = group
            if password:
                user.set_password(password)
            user.save()
            updated += 1
            continue

        if not password:
            skipped += 1
            errors.append(f"{index}-qator: yangi foydalanuvchi uchun parol kerak.")
            continue

        User.objects.create_user(
            username=username,
            password=password,
            full_name=full_name,
            workplace=workplace,
            role=normalized_role,
            group=group,
        )
        created += 1

    return Response({
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminOnly])
def import_questions_view(request):
    uploaded_file = request.FILES.get("file")
    subject_id = _to_int(request.data.get("subjectId"))
    if not uploaded_file:
        return Response({"detail": "Excel fayl yuborilmadi"}, status=status.HTTP_400_BAD_REQUEST)
    if not subject_id:
        return Response({"detail": "subjectId kerak"}, status=status.HTTP_400_BAD_REQUEST)

    subject = Subject.objects.filter(id=subject_id).first()
    if not subject:
        return Response({"detail": "Fan topilmadi"}, status=status.HTTP_404_NOT_FOUND)

    _, rows = _sheet_rows(uploaded_file)
    created = 0
    skipped = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        text = str(_get_value_by_aliases(row, ["Savollar matni", "text", "savol"])).strip()
        options = [
            str(_get_value_by_aliases(row, ["A", "a"])).strip(),
            str(_get_value_by_aliases(row, ["B", "b"])).strip(),
            str(_get_value_by_aliases(row, ["C", "S", "c", "s"])).strip(),
            str(_get_value_by_aliases(row, ["D", "d"])).strip(),
        ]
        correct_raw = _get_value_by_aliases(row, ["To'g'ri javob", "correct", "correctindex"])
        correct_index = _to_int(correct_raw, 1) - 1

        if not text or any(not option for option in options):
            skipped += 1
            errors.append(f"{index}-qator: savol yoki variantlar to'liq emas.")
            continue
        if correct_index not in {0, 1, 2, 3}:
            skipped += 1
            errors.append(f"{index}-qator: to'g'ri javob 1-4 oralig'ida bo'lishi kerak.")
            continue

        Question.objects.create(
            subject=subject,
            text=text,
            option_a=options[0],
            option_b=options[1],
            option_c=options[2],
            option_d=options[3],
            correct_index=correct_index,
        )
        created += 1

    return Response({
        "created": created,
        "skipped": skipped,
        "errors": errors[:20],
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def snapshot_view(request):
    payload = _build_snapshot_payload(request.user)
    if request.user.role == "TINGLOVCHI":
        return Response(ParticipantSnapshotSerializer(payload, context={"request": request}).data)
    return Response(SnapshotSerializer(payload, context={"request": request}).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminOnly])
def sync_snapshot_view(request):
    return Response(
        {"detail": "snapshot/sync endpoint bekor qilingan. Resource-based CRUD endpointlardan foydalaning."},
        status=status.HTTP_410_GONE,
    )
